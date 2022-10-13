from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from pathlib import Path
from openpyxl import Workbook
import openpyxl
import logging
import os
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import maskpass 


class LinkedInBot:
    def __init__(self, delay=5):
        log_fmt = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        logging.basicConfig(level=logging.INFO, format=log_fmt)
        self.delay=delay
        logging.info("Starting driver")
        self.driver_path = "./Desktop/chromedriver"
        self.delay = delay
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        self.jobs = {
            "Position": [],
            "Company": [],
            "Location" : [],
            "Link" : [],
            "Details" : []
        }

    def login(self, email, pw):
        """This function takes the users email and password to log into LinkedIn

        Args:
            email (_string_): the users email
            pw (string): the users password
        """
        logging.info("Logging in")
        self.driver.maximize_window()
        self.driver.get('https://www.linkedin.com/login')
        time.sleep(self.delay)
        self.driver.find_element(By.ID, 'username').send_keys(email)
        self.driver.find_element(By.ID, 'password').send_keys(pw)
        self.driver.find_element(By.ID, 'password').send_keys(Keys.RETURN)
        time.sleep(self.delay)

    def wait(self, t_delay=None):
        """Just easier to build this in here.
        Parameters
        ----------
        t_delay [optional] : int
            seconds to wait.
        """
        delay = self.delay if t_delay == None else t_delay
        time.sleep(delay)

    def search_linkedin(self, keywords, location):
        """Enter keywords into search bar"""
        logging.info("Searching jobs page")
        self.driver.get("https://www.linkedin.com/jobs/")
        # search based on keywords and location and hit enter
        self.wait_for_element_ready(By.CSS_SELECTOR, "input[id^='jobs-search-box-keyword-id-ember']")
        time.sleep(self.delay)
        search_bars = self.driver.find_element(By.CSS_SELECTOR, "input[id^='jobs-search-box-keyword-id-ember']")
        search_keywords = search_bars
        search_keywords.send_keys(keywords)
        time.sleep(self.delay)
        search_keywords.send_keys(Keys.RETURN)
        logging.info("Keyword search successful")
        time.sleep(self.delay) 
        
    def scroll_to(self, job_list_item):
        """Just a function that will scroll to the list item in the column 
        """
        self.driver.execute_script("arguments[0].scrollIntoView();", job_list_item)
        job_list_item.click()
        time.sleep(self.delay)   

    def get_position_data(self, job):
        """Gets the position data for a posting.
        Parameters
        ----------
        job : Selenium webelement
        Returns
        -------
        list of strings : [position, company, location, details]
        """
        try:
            [position, company, location] = job.text.split('\n')[:3]
            details = self.driver.find_element(By.ID, "job-details").text
            link = self.driver.find_element(By.CSS_SELECTOR,"a.job-card-list__title").get_attribute("href")
            return [position, company, location, details, link]
        except:
            return None
    
    def wait_for_element_ready(self, by, text):
        try:
            WebDriverWait(self.driver, self.delay).until(EC.presence_of_element_located((by, text)))
        except TimeoutException:
            logging.debug("wait_for_element_ready TimeoutException")
            pass

    def close_session(self):
        """This function closes the actual session"""
        logging.info("Closing session")
        self.driver.close()

    def create_workbook(self, path):
        workbook = xlsxwriter.Workbook(path)
        workbook.add_worksheet()
        workbook.close()

    def excel_export(self):
        data = pd.DataFrame(self.jobs, columns = ["Position", "Company", "Location", "Link", "Details"])
        workbook = openpyxl.load_workbook("../jobs.xlsx")
        writer = pd.ExcelWriter('../jobs.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        reader = pd.read_excel(r'../jobs.xlsx', engine='openpyxl')
        data.to_excel(writer,index=False,header=False, startrow=len(reader)+1)
        writer.close()

    
    def run(self, email, password, keywords, location):
        self.login(
            email=email,
            pw=password
        )
        logging.info("Begin linkedin keyword search")
        self.search_linkedin(keywords, location)
        self.wait()
        no_of_jobs = self.driver.find_element(By.CSS_SELECTOR,"div > small.jobs-search-results-list__text").get_attribute("innerText")
        # scrape pages,only do first 8 pages since after that the data isn't 
        # well suited for me anyways:  
        for page in range(2, 3):
            # get the jobs list items to scroll through:
            job_list = self.driver.find_element(By.CSS_SELECTOR, "ul.scaffold-layout__list-container")
            jobs = job_list.find_elements(By.CSS_SELECTOR, "li.jobs-search-results__list-item")
            for job in jobs:
                self.scroll_to(job)
                job_info = self.get_position_data(job)
                if job_info != None:      
                    self.jobs["Position"].append(job_info[0])
                    self.jobs["Company"].append(job_info[1])
                    self.jobs["Location"].append(job_info[2])
                    self.jobs["Link"].append(job_info[3])
                    self.jobs["Details"].append(job_info[4])
                else:
                    continue
            self.driver.find_element(By.XPATH, f"//button[@aria-label='Page {page}']").click()
            self.wait()
        logging.info("Done scraping.")
        logging.info("Closing DB connection.")
        self.close_session()


if __name__ == "__main__":
    email = input("Enter your email: ")
    pw = maskpass.askpass(prompt="Enter your password: ", mask="*")
    file = input("Enter the file name you are creating: ")
    job_keywords = input("Enter the job industry you are looking for: ")
    country = input("Enter the country: ")
    bot = LinkedInBot()
    bot.run(email, pw, job_keywords, country)
    if not os.path.exists("../"+file):
        bot.create_workbook("../"+file)
    bot.excel_export()